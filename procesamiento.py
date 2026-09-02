import csv
import io
import re
import unicodedata
import zipfile

from copy import copy
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook

from reglas_comisiones import estandarizar_equipo

VERSION_PROCESAMIENTO = "3.0-NUEVO"


HOJA_COMISIONES = "VENTAS"


ALIASES = {
    "pertenencia": [
        "PERTENENCIA",
    ],

    "afipos": [
        "AFIPOS",
    ],

    "numpos": [
        "NUMPOS",
        "NUM POS",
        "NUMERO POS",
    ],

    "afiliado": [
        "AFILIADO",
        "AFILIACION",
        "AFILIACIÓN",
        "CODIGO_AFILIADO",
        "CODIGO AFILIADO",
        "CÓDIGO AFILIADO",
        "COD AFILIADO",
        "NRO AFILIADO",
        "NUMERO AFILIADO",
    ],

    "terminal": [
        "TERMINAL",
        "NRO TERMINAL",
        "NUMERO TERMINAL",
        "NÚMERO TERMINAL",
    ],

    "serial": [
        "SERIAL",
        "SERIALES",
        "SERIAL EQUIPO",
        "SERIAL_EQUIPO",
    ],

    "monto_tx": [
        "MONTO_TRANS_BS_ACUM_MES",
        "MONTO TRANS BS ACUM MES",
        "MONTO_TRANSACCION_BS_ACUM_MES",
        "MONTO TRANSACCION BS ACUM MES",
        "MONTO_TRANS_ACUM_MES",
        "MONTO TRANS ACUM MES",
        "MONTO TX",
        "MONTO_TX",
    ],

    "mes_proceso": [
        "MES_PROCESO",
        "MES PROCESO",
    ],

    "ano_proceso": [
        "ANO_PROCESO",
        "AÑO_PROCESO",
        "ANO PROCESO",
        "AÑO PROCESO",
    ],

    "concatenar": [
        "CONCATENAR",
        "CONCATENADO",
        "AFILIADO TERMINAL",
        "AFILIADO+TERMINAL",
    ],

    "equipo": [
        "EQUIPO",
        "MODELO",
        "TIPO EQUIPO",
        "MODELO EQUIPO",
        "PRODUCTO",
    ],

    "estatus": [
        "ESTATUS",
        "STATUS",
        "ESTADO",
    ],

    "fecha_pago": [
        "FECHA DE PAGO",
        "FECHA PAGO",
    ],

    "mes_cierre": [
        "MES DE CIERRE",
    ],

    "fecha_solicitud": [
        "FECHA DE SOLICITUD RECIBIDA",
        "FECHA DE SOLICITUD  RECIBIDA",
        "FECHA SOLICITUD RECIBIDA",
    ],

    "fecha_reporte": [
        "FECHA REPORTE",
    ],

    "fecha": [
        "FECHA",
    ],

    "canal": [
        "CANAL",
    ],

    "canal_venta": [
        "CANAL DE VENTA (JORNADA QUE PERTENECE)",
        "CANAL DE VENTAS (JORNADA QUE PERTENECE)",
        "CANAL DE VENTA",
    ],

    "vendedor": [
        "VENDEDOR",
    ],

    "esquema": [
        "ESQUEMA COMERCIAL",
        "DECONTADO / FINANCIAMIENTO",
        "DE CONTADO / FINANCIAMIENTO",
    ],

    "operadora": [
        "OPERADORA",
    ],

    "banco": [
        "BANCO",
    ],

    "razon_social": [
        "RAZON SOCIAL",
        "RAZÓN SOCIAL",
    ],

    "rif": [
        "RIF",
    ],

    "telefono": [
        "TLF",
        "TELEFONO",
        "TELÉFONO",
    ],

    "direccion": [
        "DIRECCION",
        "DIRECCIÓN",
    ],

    "representante": [
        "REPRESENTANTE LEGAL",
    ],

    "correo": [
        "CORREO",
    ],

    "preafiliado": [
        "PRE AFILIADO",
        "PRE-AFILIADO",
        "PREAFILIADO",
        "PRE AFILIACION",
        "PRE-AFILIACION",
    ],

    "ag_autorizado": [
        "AG AUTORIZADO",
        "AGENTE AUTORIZADO?",
        "ES AGENTE AUTORIZADO",
    ],

    "con_tx": [
        "CON TX",
        "CON_TX",
        "ESTADO TX",
        "ESTATUS TX",
    ],

    "access": [
        "REGISTROS DE OPERADORES ACCESS",
        "REGISTRO DE OPERADORES ACCESS",
        "REGISTRO DE OPERADORES ACCESS COMERCES",
        "REGISTRO DE OPERADORES ACCESS COMMERCE",
        "ACCESS COMMERCE",
        "ACCESS COMERCE",
    ],
}


MESES_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def normalizar_texto(valor):
    if valor is None:
        return ""

    if isinstance(valor, float) and pd.isna(valor):
        return ""

    texto = str(valor).strip()
    texto = unicodedata.normalize("NFKD", texto)

    texto = "".join(
        c for c in texto
        if not unicodedata.combining(c)
    )

    texto = re.sub(r"\s+", " ", texto)
    return texto.upper()


def normalizar_nombre_columna(nombre):
    texto = normalizar_texto(nombre)

    texto = re.sub(
        r"[^A-Z0-9]+",
        " ",
        texto
    )

    return re.sub(
        r"\s+",
        " ",
        texto
    ).strip()


def buscar_columnas(df, tipo):
    if df is None:
        return []

    aliases = {
        normalizar_nombre_columna(alias)
        for alias in ALIASES.get(tipo, [])
    }

    encontrados = []

    for col in df.columns:
        normalizada = normalizar_nombre_columna(col)

        if normalizada in aliases:
            encontrados.append(col)

    if encontrados:
        return encontrados

    for col in df.columns:
        normalizada = normalizar_nombre_columna(col)

        for alias in aliases:
            if alias and (
                alias in normalizada
                or normalizada in alias
            ):
                encontrados.append(col)
                break

    return encontrados


def buscar_columna(df, tipo, preferir_ultima=False):
    columnas = buscar_columnas(
        df,
        tipo
    )

    if not columnas:
        return None

    if preferir_ultima:
        return columnas[-1]

    return columnas[0]


def normalizar_identificador(valor):
    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    texto = str(valor).strip()

    if normalizar_texto(texto) in {
        "",
        "N/A",
        "N/D",
        "NA",
        "ND",
        "NAN",
        "NONE",
        "-",
    }:
        return normalizar_texto(texto)

    if re.fullmatch(
        r"[+-]?\d+\.0+",
        texto
    ):
        texto = texto.split(".")[0]

    texto = re.sub(
        r"\s+",
        "",
        texto
    )

    return texto


def normalizar_numero(valor):
    if valor is None:
        return None

    try:
        if pd.isna(valor):
            return None
    except Exception:
        pass

    texto = str(valor).strip()

    if normalizar_texto(texto) in {
        "",
        "N/A",
        "N/D",
        "NA",
        "ND",
        "NAN",
        "-",
    }:
        return None

    texto = re.sub(
        r"[^0-9,\.\-]",
        "",
        texto
    )

    if not texto:
        return None

    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = (
                texto
                .replace(".", "")
                .replace(",", ".")
            )
        else:
            texto = texto.replace(",", "")

    elif "," in texto:
        texto = texto.replace(",", ".")

    try:
        return float(texto)

    except ValueError:
        return None


def crear_concatenar(afiliado, terminal):
    """
    El CONCATENAR real del archivo es AFILIADO + TERMINAL,
    sin separador.
    Ejemplo: 87818535 + 1 = 878185351
    """
    afiliado = normalizar_identificador(
        afiliado
    )

    terminal = normalizar_identificador(
        terminal
    )

    if (
        not afiliado
        or afiliado in {"N/A", "N/D", "NA", "ND", "-"}
        or not terminal
        or terminal in {"N/A", "N/D", "NA", "ND", "-"}
    ):
        return ""

    return f"{afiliado}{terminal}"


def valor_es_si(valor):
    return normalizar_texto(valor) in {
        "SI",
        "SÍ",
        "YES",
        "Y",
        "1",
        "TRUE",
        "X",
    }


def viernes_semana_actual():
    hoy = datetime.now(
        ZoneInfo("America/Caracas")
    ).date()

    return hoy + timedelta(
        days=(4 - hoy.weekday())
    )


def nombre_mes_cierre(fecha):
    fecha = pd.to_datetime(
        fecha,
        errors="coerce"
    )

    if pd.isna(fecha):
        return ""

    mes = MESES_ES.get(
        int(fecha.month),
        ""
    )

    return f"{mes} ({int(fecha.year)})"


def leer_excel_general(archivo):
    archivo.seek(0)

    return pd.read_excel(
        archivo,
        dtype=object,
        engine="openpyxl",
    )


def obtener_hoja_comisiones(archivo):
    archivo.seek(0)

    excel = pd.ExcelFile(
        archivo,
        engine="openpyxl"
    )

    if HOJA_COMISIONES in excel.sheet_names:
        return HOJA_COMISIONES

    # Esto permite probar con una muestra que tenga una sola hoja,
    # sin cambiar la regla del archivo real.
    if len(excel.sheet_names) == 1:
        return excel.sheet_names[0]

    raise ValueError(
        "No encontré la hoja 'VENTAS' "
        "en el archivo de Comisiones."
    )


def leer_excel_comisiones(archivo):
    hoja = obtener_hoja_comisiones(
        archivo
    )

    archivo.seek(0)

    df = pd.read_excel(
        archivo,
        sheet_name=hoja,
        dtype=object,
        engine="openpyxl",
    )

    df.attrs[
        "hoja_origen"
    ] = hoja

    return df


def detectar_csv(stream):
    posicion = stream.tell()
    muestra = stream.read(65536)
    stream.seek(posicion)

    encoding = "utf-8-sig"

    try:
        texto = muestra.decode(
            encoding
        )

    except UnicodeDecodeError:
        encoding = "latin-1"

        texto = muestra.decode(
            encoding,
            errors="replace"
        )

    try:
        dialecto = csv.Sniffer().sniff(
            texto,
            delimiters=[
                ",",
                ";",
                "\t",
                "|",
            ]
        )

        separador = dialecto.delimiter

    except csv.Error:
        separador = (
            ";"
            if texto.count(";") > texto.count(",")
            else ","
        )

    return encoding, separador


def extraer_serial_terminal_r34(valor):
    """
    El campo TERMINAL del R34 puede venir así:
    000181252538805 [00995] CASTLES DYNAMO...
    Conservamos el primer bloque numérico largo.
    """
    texto = str(
        valor
        if valor is not None
        else ""
    ).strip()

    match = re.search(
        r"\d{8,}",
        texto
    )

    if match:
        serial = match.group(0)

        # Quitar ceros de relleno a la izquierda
        # solo si realmente existen.
        serial_limpio = serial.lstrip("0")

        return serial_limpio or "0"

    return ""


def procesar_csv_r34(
    stream,
    nombre,
    chunksize=100_000
):
    stream.seek(0)

    encoding, separador = detectar_csv(
        stream
    )

    stream.seek(0)

    partes = []
    filas_leidas = 0
    filas_credicardpos = 0
    columnas = None
    mes_proceso_detectado = None
    ano_proceso_detectado = None

    lector = pd.read_csv(
        stream,
        sep=separador,
        encoding=encoding,
        dtype=str,
        chunksize=chunksize,
        on_bad_lines="skip",
        low_memory=False,
    )

    for chunk in lector:
        filas_leidas += len(chunk)

        if columnas is None:
            columnas = {
                "pertenencia":
                    buscar_columna(
                        chunk,
                        "pertenencia"
                    ),

                "afipos":
                    buscar_columna(
                        chunk,
                        "afipos"
                    ),

                "afiliado":
                    buscar_columna(
                        chunk,
                        "afiliado"
                    ),

                "numpos":
                    buscar_columna(
                        chunk,
                        "numpos"
                    ),

                "terminal":
                    buscar_columna(
                        chunk,
                        "terminal"
                    ),

                "serial":
                    buscar_columna(
                        chunk,
                        "serial"
                    ),

                "monto_tx":
                    buscar_columna(
                        chunk,
                        "monto_tx"
                    ),

                "mes_proceso":
                    buscar_columna(
                        chunk,
                        "mes_proceso"
                    ),

                "ano_proceso":
                    buscar_columna(
                        chunk,
                        "ano_proceso"
                    ),
            }

            if columnas["pertenencia"] is None:
                raise ValueError(
                    f"No se encontró la columna "
                    f"PERTENENCIA en {nombre}."
                )

            if columnas["afipos"] is None:
                raise ValueError(
                    f"No se encontró la columna "
                    f"AFIPOS en {nombre}."
                )

            if columnas["monto_tx"] is None:
                raise ValueError(
                    f"No se encontró MONTO_TRANS_BS_ACUM_MES "
                    f"en {nombre}."
                )

        pertenencia = (
            chunk[
                columnas["pertenencia"]
            ]
            .map(normalizar_texto)
        )

        # El R34 real contiene valores como:
        # CREDICARDPOS CDM
        # CREDICARDPOS OCCIDENT CDM
        # Por eso NO debe compararse con igualdad exacta.
        mascara = pertenencia.str.contains(
            "CREDICARDPOS",
            na=False,
        )

        filtrado = chunk.loc[
            mascara
        ].copy()

        filas_credicardpos += len(
            filtrado
        )

        if filtrado.empty:
            continue

        filtrado[
            "__CONCATENAR"
        ] = filtrado[
            columnas["afipos"]
        ].map(
            normalizar_identificador
        )

        if columnas["afiliado"]:
            filtrado[
                "__AFILIADO"
            ] = filtrado[
                columnas["afiliado"]
            ].map(
                normalizar_identificador
            )
        else:
            filtrado[
                "__AFILIADO"
            ] = ""

        if columnas["numpos"]:
            filtrado[
                "__TERMINAL_NUM"
            ] = filtrado[
                columnas["numpos"]
            ].map(
                normalizar_identificador
            )
        else:
            filtrado[
                "__TERMINAL_NUM"
            ] = ""

        if columnas["serial"]:
            filtrado[
                "__SERIAL_R34"
            ] = filtrado[
                columnas["serial"]
            ].map(
                normalizar_identificador
            )
        else:
            filtrado[
                "__SERIAL_R34"
            ] = ""

        if columnas["terminal"]:
            filtrado[
                "__SERIAL_TERMINAL_R34"
            ] = filtrado[
                columnas["terminal"]
            ].map(
                extraer_serial_terminal_r34
            )
        else:
            filtrado[
                "__SERIAL_TERMINAL_R34"
            ] = ""

        filtrado[
            "__MONTO_TX"
        ] = filtrado[
            columnas["monto_tx"]
        ].map(
            normalizar_numero
        )

        if (
            mes_proceso_detectado is None
            and columnas["mes_proceso"]
        ):
            serie = pd.to_numeric(
                filtrado[
                    columnas["mes_proceso"]
                ],
                errors="coerce"
            ).dropna()

            if not serie.empty:
                mes_proceso_detectado = int(
                    serie.iloc[0]
                )

        if (
            ano_proceso_detectado is None
            and columnas["ano_proceso"]
        ):
            serie = pd.to_numeric(
                filtrado[
                    columnas["ano_proceso"]
                ],
                errors="coerce"
            ).dropna()

            if not serie.empty:
                ano_proceso_detectado = int(
                    serie.iloc[0]
                )

        partes.append(
            filtrado[
                [
                    "__CONCATENAR",
                    "__AFILIADO",
                    "__TERMINAL_NUM",
                    "__SERIAL_R34",
                    "__SERIAL_TERMINAL_R34",
                    "__MONTO_TX",
                ]
            ]
        )

    if partes:
        resultado = pd.concat(
            partes,
            ignore_index=True
        )
    else:
        resultado = pd.DataFrame(
            columns=[
                "__CONCATENAR",
                "__AFILIADO",
                "__TERMINAL_NUM",
                "__SERIAL_R34",
                "__SERIAL_TERMINAL_R34",
                "__MONTO_TX",
            ]
        )

    detalle = {
        "archivo": nombre,
        "filas_leidas": filas_leidas,
        "filas_credicardpos":
            filas_credicardpos,
        "mes_proceso":
            mes_proceso_detectado,
        "ano_proceso":
            ano_proceso_detectado,
    }

    return resultado, detalle


def procesar_r34(
    archivos_r34,
    chunksize=100_000
):
    partes = []
    detalles = []

    for archivo in archivos_r34:
        archivo.seek(0)

        nombre = archivo.name.lower()

        if nombre.endswith(".zip"):
            datos = archivo.getvalue()

            with zipfile.ZipFile(
                io.BytesIO(datos)
            ) as zf:
                csvs = [
                    nombre_interno
                    for nombre_interno
                    in zf.namelist()
                    if (
                        nombre_interno.lower().endswith(".csv")
                        and not nombre_interno.startswith("__MACOSX/")
                    )
                ]

                if not csvs:
                    raise ValueError(
                        f"{archivo.name} no contiene ningún CSV."
                    )

                for csv_interno in csvs:
                    with zf.open(
                        csv_interno,
                        "r"
                    ) as stream:
                        df, detalle = procesar_csv_r34(
                            stream,
                            f"{archivo.name}::{csv_interno}",
                            chunksize
                        )

                        partes.append(df)
                        detalles.append(detalle)

        elif nombre.endswith(".csv"):
            df, detalle = procesar_csv_r34(
                archivo,
                archivo.name,
                chunksize
            )

            partes.append(df)
            detalles.append(detalle)

        else:
            raise ValueError(
                f"Formato de R34 no soportado: "
                f"{archivo.name}"
            )

    r34 = (
        pd.concat(
            partes,
            ignore_index=True
        )
        if partes
        else pd.DataFrame()
    )

    return r34, detalles


def preparar_comisiones(
    archivo_comisiones
):
    df = leer_excel_comisiones(
        archivo_comisiones
    )

    df = df.dropna(
        how="all"
    ).copy()

    df["__ORIGEN"] = "COMISIONES"

    # En el ejemplo real existen columnas repetidas
    # CONCATENAR y SERIAL. Para la lógica principal
    # preferimos la última, que corresponde al bloque
    # principal de la hoja VENTAS.
    col_afiliado = buscar_columna(
        df,
        "afiliado",
        preferir_ultima=True
    )

    col_terminal = buscar_columna(
        df,
        "terminal",
        preferir_ultima=True
    )

    col_serial = buscar_columna(
        df,
        "serial",
        preferir_ultima=True
    )

    col_equipo = buscar_columna(
        df,
        "equipo",
        preferir_ultima=True
    )

    col_estatus = buscar_columna(
        df,
        "estatus",
        preferir_ultima=True
    )

    if col_afiliado is None:
        raise ValueError(
            "En la hoja VENTAS no encontré "
            "la columna AFILIADO."
        )

    if col_terminal is None:
        raise ValueError(
            "En la hoja VENTAS no encontré "
            "la columna TERMINAL."
        )

    df[
        "__AFILIADO"
    ] = df[
        col_afiliado
    ].map(
        normalizar_identificador
    )

    df[
        "__TERMINAL"
    ] = df[
        col_terminal
    ].map(
        normalizar_identificador
    )

    if col_serial:
        df[
            "__SERIAL_COMISION"
        ] = df[
            col_serial
        ].map(
            normalizar_identificador
        )
    else:
        df[
            "__SERIAL_COMISION"
        ] = ""

    if col_equipo:
        df[
            "__EQUIPO_STD"
        ] = df[
            col_equipo
        ].map(
            estandarizar_equipo
        )
    else:
        df[
            "__EQUIPO_STD"
        ] = ""

    df[
        "__CONCATENAR"
    ] = [
        crear_concatenar(
            afiliado,
            terminal
        )
        for afiliado, terminal
        in zip(
            df["__AFILIADO"],
            df["__TERMINAL"]
        )
    ]

    if col_estatus:
        df[
            "__ESTATUS_NORMALIZADO"
        ] = df[
            col_estatus
        ].map(
            normalizar_texto
        )
    else:
        df[
            "__ESTATUS_NORMALIZADO"
        ] = ""

    df["__ROW_ID"] = range(
        1,
        len(df) + 1
    )

    return df


def preparar_ventas(
    archivos_ventas
):
    partes = []
    excluidas = []
    advertencias = []

    for archivo in archivos_ventas:
        df = leer_excel_general(
            archivo
        )

        df = df.dropna(
            how="all"
        ).copy()

        df[
            "__ARCHIVO_ORIGEN"
        ] = archivo.name

        col_afiliado = buscar_columna(
            df,
            "afiliado"
        )

        col_terminal = buscar_columna(
            df,
            "terminal"
        )

        col_equipo = buscar_columna(
            df,
            "equipo"
        )

        col_pre = buscar_columna(
            df,
            "preafiliado"
        )

        col_ag = buscar_columna(
            df,
            "ag_autorizado"
        )

        if col_afiliado is None:
            raise ValueError(
                f"No encontré AFILIADO "
                f"en {archivo.name}."
            )

        if col_terminal is None:
            raise ValueError(
                f"No encontré TERMINAL "
                f"en {archivo.name}."
            )

        df[
            "__AFILIADO"
        ] = df[
            col_afiliado
        ].map(
            normalizar_identificador
        )

        df[
            "__TERMINAL"
        ] = df[
            col_terminal
        ].map(
            normalizar_identificador
        )

        df[
            "__CONCATENAR"
        ] = [
            crear_concatenar(
                afiliado,
                terminal
            )
            for afiliado, terminal
            in zip(
                df["__AFILIADO"],
                df["__TERMINAL"]
            )
        ]

        if col_equipo:
            df[
                "__EQUIPO_STD"
            ] = df[
                col_equipo
            ].map(
                estandarizar_equipo
            )
        else:
            df[
                "__EQUIPO_STD"
            ] = ""

        df[
            "__MOTIVO_EXCLUSION"
        ] = ""

        mascara_terminal_0 = (
            df[
                "__TERMINAL"
            ].isin(
                {
                    "0",
                    "0.0",
                }
            )
        )

        df.loc[
            mascara_terminal_0,
            "__MOTIVO_EXCLUSION"
        ] = "Terminal = 0 / No Simcard"

        if col_pre:
            mascara_pre = df[
                col_pre
            ].map(
                lambda valor:
                valor_es_si(valor)
                or normalizar_texto(valor).startswith("PRE")
            )

            df.loc[
                mascara_pre
                & df[
                    "__MOTIVO_EXCLUSION"
                ].eq(""),
                "__MOTIVO_EXCLUSION"
            ] = "Pre-afiliado"

        # La guía dice que las ventas de AG Autorizados
        # deben excluirse, pero el reporte de ejemplo no
        # trae una columna explícita con ese indicador.
        # No inventamos una condición.
        if col_ag:
            mascara_ag = df[
                col_ag
            ].map(
                valor_es_si
            )

            df.loc[
                mascara_ag
                & df[
                    "__MOTIVO_EXCLUSION"
                ].eq(""),
                "__MOTIVO_EXCLUSION"
            ] = "AG Autorizado"
        else:
            advertencias.append(
                f"{archivo.name}: el reporte no trae "
                f"una columna explícita para identificar "
                f"AG Autorizados; esa exclusión no se "
                f"aplicó automáticamente."
            )

        excluidas.append(
            df[
                df[
                    "__MOTIVO_EXCLUSION"
                ].ne("")
            ].copy()
        )

        partes.append(
            df[
                df[
                    "__MOTIVO_EXCLUSION"
                ].eq("")
            ].copy()
        )

    ventas = (
        pd.concat(
            partes,
            ignore_index=True
        )
        if partes
        else pd.DataFrame()
    )

    ventas_excluidas = (
        pd.concat(
            excluidas,
            ignore_index=True
        )
        if excluidas
        else pd.DataFrame()
    )

    if not ventas.empty:
        mascara_dup = (
            ventas[
                "__CONCATENAR"
            ].ne("")
            & ventas.duplicated(
                "__CONCATENAR",
                keep="first"
            )
        )

        ventas_duplicadas = ventas[
            mascara_dup
        ].copy()

        ventas = ventas[
            ~mascara_dup
        ].copy()

    else:
        ventas_duplicadas = pd.DataFrame()

    return (
        ventas,
        ventas_excluidas,
        ventas_duplicadas,
        advertencias,
    )


def asignar_si_existe(
    destino,
    fila_indice,
    df_destino,
    df_origen,
    tipo_destino,
    tipo_origen=None,
    transformacion=None,
):
    if tipo_origen is None:
        tipo_origen = tipo_destino

    columnas_destino = buscar_columnas(
        df_destino,
        tipo_destino
    )

    columna_origen = buscar_columna(
        df_origen,
        tipo_origen
    )

    if (
        not columnas_destino
        or columna_origen is None
    ):
        return

    valores = df_origen[
        columna_origen
    ].tolist()

    if transformacion:
        valores = [
            transformacion(valor)
            for valor in valores
        ]

    for columna_destino in columnas_destino:
        destino[
            columna_destino
        ] = valores


def crear_filas_nuevas(
    ventas_nuevas,
    comisiones
):
    if ventas_nuevas.empty:
        return pd.DataFrame(
            columns=comisiones.columns
        )

    nuevas = pd.DataFrame(
        index=range(
            len(ventas_nuevas)
        ),
        columns=comisiones.columns,
        dtype=object,
    )

    # Primero copiamos columnas que tengan
    # exactamente el mismo nombre normalizado.
    for columna_destino in comisiones.columns:
        if str(
            columna_destino
        ).startswith("__"):
            continue

        destino_norm = normalizar_nombre_columna(
            columna_destino
        )

        for columna_origen in ventas_nuevas.columns:
            if str(
                columna_origen
            ).startswith("__"):
                continue

            origen_norm = normalizar_nombre_columna(
                columna_origen
            )

            # pandas agrega .1 a nombres duplicados.
            origen_norm = re.sub(
                r"\s+\d+$",
                "",
                origen_norm
            )

            destino_base = re.sub(
                r"\s+\d+$",
                "",
                destino_norm
            )

            if destino_base == origen_norm:
                nuevas[
                    columna_destino
                ] = ventas_nuevas[
                    columna_origen
                ].values

                break

    # Mapeos de columnas equivalentes
    mapeos = [
        ("afiliado", "afiliado"),
        ("terminal", "terminal"),
        ("serial", "serial"),
        ("equipo", "equipo"),
        ("canal", "canal"),
        ("canal_venta", "canal_venta"),
        ("vendedor", "vendedor"),
        ("operadora", "operadora"),
        ("banco", "banco"),
        ("razon_social", "razon_social"),
        ("rif", "rif"),
        ("telefono", "telefono"),
        ("direccion", "direccion"),
        ("representante", "representante"),
        ("correo", "correo"),
    ]

    for tipo_destino, tipo_origen in mapeos:
        columnas_destino = buscar_columnas(
            comisiones,
            tipo_destino
        )

        columna_origen = buscar_columna(
            ventas_nuevas,
            tipo_origen
        )

        if (
            not columnas_destino
            or columna_origen is None
        ):
            continue

        valores = ventas_nuevas[
            columna_origen
        ].tolist()

        for columna_destino in columnas_destino:
            nuevas[
                columna_destino
            ] = valores

    # CONCATENAR se calcula, no se confía en la fórmula
    # del reporte.
    for columna in buscar_columnas(
        comisiones,
        "concatenar"
    ):
        nuevas[
            columna
        ] = ventas_nuevas[
            "__CONCATENAR"
        ].values

    # Serial auxiliar y serial principal
    for columna in buscar_columnas(
        comisiones,
        "serial"
    ):
        col_serial_ventas = buscar_columna(
            ventas_nuevas,
            "serial"
        )

        if col_serial_ventas:
            nuevas[
                columna
            ] = ventas_nuevas[
                col_serial_ventas
            ].values

    # Estatus
    for columna in buscar_columnas(
        comisiones,
        "estatus"
    ):
        nuevas[
            columna
        ] = "Pendiente"

    # Esquema comercial
    columnas_esquema = buscar_columnas(
        comisiones,
        "esquema"
    )

    origen_esquema = buscar_columna(
        ventas_nuevas,
        "esquema"
    )

    if (
        columnas_esquema
        and origen_esquema
    ):
        for columna in columnas_esquema:
            nuevas[
                columna
            ] = ventas_nuevas[
                origen_esquema
            ].values

    # MES DE CIERRE desde FECHA REPORTE
    columnas_mes = buscar_columnas(
        comisiones,
        "mes_cierre"
    )

    fecha_reporte = buscar_columna(
        ventas_nuevas,
        "fecha_reporte"
    )

    if (
        columnas_mes
        and fecha_reporte
    ):
        valores = ventas_nuevas[
            fecha_reporte
        ].map(
            nombre_mes_cierre
        ).values

        for columna in columnas_mes:
            nuevas[
                columna
            ] = valores

    # FECHA DE SOLICITUD RECIBIDA desde FECHA de venta
    columnas_solicitud = buscar_columnas(
        comisiones,
        "fecha_solicitud"
    )

    fecha_venta = buscar_columna(
        ventas_nuevas,
        "fecha"
    )

    if (
        columnas_solicitud
        and fecha_venta
    ):
        for columna in columnas_solicitud:
            nuevas[
                columna
            ] = ventas_nuevas[
                fecha_venta
            ].values

    nuevas[
        "__ORIGEN"
    ] = "VENTAS_NUEVAS"

    nuevas[
        "__AFILIADO"
    ] = ventas_nuevas[
        "__AFILIADO"
    ].values

    nuevas[
        "__TERMINAL"
    ] = ventas_nuevas[
        "__TERMINAL"
    ].values

    nuevas[
        "__CONCATENAR"
    ] = ventas_nuevas[
        "__CONCATENAR"
    ].values

    nuevas[
        "__EQUIPO_STD"
    ] = ventas_nuevas[
        "__EQUIPO_STD"
    ].values

    col_serial_ventas = buscar_columna(
        ventas_nuevas,
        "serial"
    )

    if col_serial_ventas:
        nuevas[
            "__SERIAL_COMISION"
        ] = ventas_nuevas[
            col_serial_ventas
        ].map(
            normalizar_identificador
        ).values
    else:
        nuevas[
            "__SERIAL_COMISION"
        ] = ""

    nuevas[
        "__ESTATUS_NORMALIZADO"
    ] = "PENDIENTE"

    return nuevas


def integrar_ventas(
    comisiones,
    ventas
):
    if ventas.empty:
        return (
            comisiones.copy(),
            pd.DataFrame(),
            pd.DataFrame(),
        )

    claves_existentes = set(
        comisiones.loc[
            comisiones[
                "__CONCATENAR"
            ].ne(""),
            "__CONCATENAR"
        ].astype(str)
    )

    mascara_nueva = (
        ventas[
            "__CONCATENAR"
        ].ne("")
        & ~ventas[
            "__CONCATENAR"
        ].isin(
            claves_existentes
        )
    )

    ventas_nuevas = ventas[
        mascara_nueva
    ].copy()

    ventas_existentes = ventas[
        ~mascara_nueva
    ].copy()

    filas_nuevas = crear_filas_nuevas(
        ventas_nuevas,
        comisiones
    )

    combinado = pd.concat(
        [
            comisiones,
            filas_nuevas,
        ],
        ignore_index=True,
        sort=False,
    )

    combinado[
        "__ROW_ID"
    ] = range(
        1,
        len(combinado) + 1
    )

    return (
        combinado,
        ventas_nuevas,
        ventas_existentes,
    )


def preparar_access(
    archivo_access
):
    df = leer_excel_general(
        archivo_access
    )

    col_afiliado = buscar_columna(
        df,
        "afiliado"
    )

    if col_afiliado is None:
        raise ValueError(
            "No encontré la columna AFILIADO "
            "en Access Commerce."
        )

    df[
        "__AFILIADO"
    ] = df[
        col_afiliado
    ].map(
        normalizar_identificador
    )

    afiliados = set(
        df.loc[
            df[
                "__AFILIADO"
            ].ne(""),
            "__AFILIADO"
        ].astype(str)
    )

    return df, afiliados


def crear_lookup_r34(r34):
    lookup = {}

    if r34.empty:
        return lookup

    for _, fila in r34.iterrows():
        clave = normalizar_identificador(
            fila.get(
                "__CONCATENAR",
                ""
            )
        )

        if not clave:
            continue

        actual = lookup.get(
            clave
        )

        if actual is None:
            lookup[
                clave
            ] = fila.to_dict()

            continue

        monto_nuevo = fila.get(
            "__MONTO_TX"
        )

        monto_actual = actual.get(
            "__MONTO_TX"
        )

        if (
            monto_nuevo is not None
            and not pd.isna(
                monto_nuevo
            )
        ):
            if (
                monto_actual is None
                or pd.isna(
                    monto_actual
                )
                or float(
                    monto_nuevo
                )
                > float(
                    monto_actual
                )
            ):
                lookup[
                    clave
                ] = fila.to_dict()

    return lookup


def estado_transaccion(monto):
    if (
        monto is None
        or pd.isna(monto)
    ):
        return "N/A"

    monto = float(monto)

    if monto > 1000:
        return "CON_TX"

    if 0.01 <= monto <= 999.99:
        return "C/P SIN TX"

    if monto == 0:
        return "SIN TX"

    if monto == 1000:
        return "REVISAR 1000"

    return "REVISAR"


def columna_monto_tx_mes(
    df,
    mes_numero
):
    if not mes_numero:
        return None

    nombre_mes = MESES_ES.get(
        int(mes_numero)
    )

    if not nombre_mes:
        return None

    for col in df.columns:
        norm = normalizar_nombre_columna(
            col
        )

        if (
            norm.startswith("MONTO TX")
            and nombre_mes in norm
        ):
            return col

    return None


def recalcular_comisiones(
    df,
    r34,
    afiliados_access,
    mes_r34=None,
):
    resultado = df.copy()

    lookup = crear_lookup_r34(
        r34
    )

    col_afiliado = buscar_columna(
        resultado,
        "afiliado",
        preferir_ultima=True
    )

    col_terminal = buscar_columna(
        resultado,
        "terminal",
        preferir_ultima=True
    )

    col_serial = buscar_columna(
        resultado,
        "serial",
        preferir_ultima=True
    )

    col_equipo = buscar_columna(
        resultado,
        "equipo",
        preferir_ultima=True
    )

    col_estatus = buscar_columna(
        resultado,
        "estatus",
        preferir_ultima=True
    )

    col_fecha_pago = buscar_columna(
        resultado,
        "fecha_pago",
        preferir_ultima=True
    )

    col_tx = buscar_columna(
        resultado,
        "con_tx",
        preferir_ultima=True
    )

    col_access = buscar_columna(
        resultado,
        "access",
        preferir_ultima=True
    )

    col_monto_mes = columna_monto_tx_mes(
        resultado,
        mes_r34
    )

    if col_afiliado:
        resultado[
            "__AFILIADO"
        ] = resultado[
            col_afiliado
        ].map(
            normalizar_identificador
        )

    if col_terminal:
        resultado[
            "__TERMINAL"
        ] = resultado[
            col_terminal
        ].map(
            normalizar_identificador
        )

    if col_serial:
        resultado[
            "__SERIAL_COMISION"
        ] = resultado[
            col_serial
        ].map(
            normalizar_identificador
        )

    if col_equipo:
        resultado[
            "__EQUIPO_STD"
        ] = resultado[
            col_equipo
        ].map(
            estandarizar_equipo
        )

    resultado[
        "__CONCATENAR"
    ] = [
        crear_concatenar(
            afiliado,
            terminal
        )
        for afiliado, terminal
        in zip(
            resultado["__AFILIADO"],
            resultado["__TERMINAL"]
        )
    ]

    if col_estatus:
        resultado[
            "__ESTATUS_NORMALIZADO"
        ] = resultado[
            col_estatus
        ].map(
            normalizar_texto
        )
    else:
        resultado[
            "__ESTATUS_NORMALIZADO"
        ] = ""

    duplicado = (
        resultado[
            "__CONCATENAR"
        ].ne("")
        & resultado.duplicated(
            "__CONCATENAR",
            keep=False
        )
    )

    montos = []
    seriales_r34 = []
    estados_tx = []
    access_lista = []
    aplica_lista = []
    motivos = []

    for idx, fila in resultado.iterrows():
        estatus = normalizar_texto(
            fila.get(
                "__ESTATUS_NORMALIZADO",
                ""
            )
        )

        # La guía indica que el trabajo de validación
        # comienza con Estatus pendiente.
        # No reabrimos PAGADO ni DESINSTALADO.
        es_pendiente = (
            estatus == "PENDIENTE"
            or fila.get("__ORIGEN") == "VENTAS_NUEVAS"
        )

        if not es_pendiente:
            montos.append(
                fila.get(
                    col_monto_mes,
                    None
                )
                if col_monto_mes
                else None
            )

            seriales_r34.append("")
            estados_tx.append(
                fila.get(
                    col_tx,
                    ""
                )
                if col_tx
                else ""
            )

            access_lista.append(
                fila.get(
                    col_access,
                    ""
                )
                if col_access
                else ""
            )

            aplica_lista.append("")
            motivos.append("")
            continue

        clave = normalizar_identificador(
            fila.get(
                "__CONCATENAR",
                ""
            )
        )

        equipo = estandarizar_equipo(
            fila.get(
                "__EQUIPO_STD",
                ""
            )
        )

        serial_comision = (
            normalizar_identificador(
                fila.get(
                    "__SERIAL_COMISION",
                    ""
                )
            )
        )

        razones = []

        if (
            clave
            and duplicado.loc[idx]
        ):
            razones.append(
                "Duplicado afiliado+terminal"
            )

        if not clave:
            razones.append(
                "Falta afiliado o terminal"
            )

        registro_r34 = lookup.get(
            clave
        )

        if registro_r34 is None:
            monto = None
            serial_r34 = ""

            razones.append(
                "No encontrado en R34"
            )

        else:
            monto = registro_r34.get(
                "__MONTO_TX"
            )

            if equipo == "Pinpagos":
                serial_r34 = (
                    registro_r34.get(
                        "__SERIAL_TERMINAL_R34",
                        ""
                    )
                    or registro_r34.get(
                        "__SERIAL_R34",
                        ""
                    )
                )
            else:
                serial_r34 = (
                    registro_r34.get(
                        "__SERIAL_R34",
                        ""
                    )
                )

        serial_r34 = normalizar_identificador(
            serial_r34
        )

        if serial_comision in {
            "N/A",
            "N/D",
            "NA",
            "ND",
        }:
            razones.append(
                f"Serial {serial_comision} "
                f"requiere revisión AS400"
            )

        elif (
            serial_comision
            and serial_r34
            and serial_comision != serial_r34
        ):
            razones.append(
                "Serial no coincide con R34"
            )

        estado_tx = estado_transaccion(
            monto
        )

        if equipo == "Pinpagos":
            access = "NO APLICA"
            aplica = (
                estado_tx == "CON_TX"
            )

        else:
            afiliado = normalizar_identificador(
                fila.get(
                    "__AFILIADO",
                    ""
                )
            )

            access = (
                "SI"
                if afiliado in afiliados_access
                else "NO"
            )

            aplica = (
                estado_tx == "CON_TX"
                and access == "SI"
            )

        if estado_tx in {
            "N/A",
            "REVISAR",
            "REVISAR 1000",
        }:
            razones.append(
                f"Estado TX: {estado_tx}"
            )

        montos.append(monto)
        seriales_r34.append(serial_r34)
        estados_tx.append(estado_tx)
        access_lista.append(access)
        aplica_lista.append(
            "SI"
            if aplica
            else "NO"
        )

        motivos.append(
            " | ".join(
                dict.fromkeys(
                    razones
                )
            )
        )

        if col_tx:
            resultado.loc[
                idx,
                col_tx
            ] = estado_tx

        if col_access:
            resultado.loc[
                idx,
                col_access
            ] = access

        if col_monto_mes:
            resultado.loc[
                idx,
                col_monto_mes
            ] = monto

        if (
            aplica
            and col_estatus
        ):
            resultado.loc[
                idx,
                col_estatus
            ] = "Aplica Pago"

            resultado.loc[
                idx,
                "__ESTATUS_NORMALIZADO"
            ] = "APLICA PAGO"

            if col_fecha_pago:
                resultado.loc[
                    idx,
                    col_fecha_pago
                ] = viernes_semana_actual()

    resultado[
        "__MONTO_TX_R34"
    ] = montos

    resultado[
        "__SERIAL_R34_COMPARADO"
    ] = seriales_r34

    resultado[
        "__ESTADO_TX_CALCULADO"
    ] = estados_tx

    resultado[
        "__ACCESS_CALCULADO"
    ] = access_lista

    resultado[
        "__APLICA_PAGO_CALCULADO"
    ] = aplica_lista

    resultado[
        "__MOTIVO_REVISION"
    ] = motivos

    resultado[
        "__REQUIERE_REVISION"
    ] = (
        resultado[
            "__MOTIVO_REVISION"
        ]
        .astype(str)
        .str.len()
        > 0
    )

    return resultado


def determinar_mes_r34(
    detalle_r34
):
    for detalle in detalle_r34:
        mes = detalle.get(
            "mes_proceso"
        )

        if mes:
            return int(mes)

    return None


def procesar_todo(
    archivos_r34,
    archivos_ventas,
    archivo_comisiones,
    archivo_access,
    chunksize=100_000,
):
    archivo_comisiones.seek(0)

    bytes_comisiones_original = (
        archivo_comisiones.getvalue()
    )

    hoja_comisiones = obtener_hoja_comisiones(
        archivo_comisiones
    )

    r34, detalle_r34 = procesar_r34(
        archivos_r34,
        chunksize
    )

    comisiones = preparar_comisiones(
        archivo_comisiones
    )

    cantidad_original = len(
        comisiones
    )

    (
        ventas_limpias,
        ventas_excluidas,
        ventas_duplicadas,
        advertencias,
    ) = preparar_ventas(
        archivos_ventas
    )

    (
        combinado,
        ventas_nuevas,
        ventas_existentes,
    ) = integrar_ventas(
        comisiones,
        ventas_limpias
    )

    (
        access_df,
        afiliados_access,
    ) = preparar_access(
        archivo_access
    )

    mes_r34 = determinar_mes_r34(
        detalle_r34
    )

    final = recalcular_comisiones(
        combinado,
        r34,
        afiliados_access,
        mes_r34=mes_r34,
    )

    return {
        "final":
            final,

        "r34":
            r34,

        "detalle_r34":
            detalle_r34,

        "ventas_limpias":
            ventas_limpias,

        "ventas_nuevas":
            ventas_nuevas,

        "ventas_existentes":
            ventas_existentes,

        "ventas_excluidas":
            ventas_excluidas,

        "ventas_duplicadas":
            ventas_duplicadas,

        "advertencias":
            advertencias,

        "afiliados_access":
            afiliados_access,

        "cantidad_original":
            cantidad_original,

        "bytes_comisiones_original":
            bytes_comisiones_original,

        "hoja_comisiones":
            hoja_comisiones,

        "mes_r34":
            mes_r34,
    }


def copiar_estilo_fila(
    ws,
    fila_origen,
    fila_destino,
):
    if fila_origen < 1:
        return

    for col in range(
        1,
        ws.max_column + 1
    ):
        origen = ws.cell(
            row=fila_origen,
            column=col
        )

        destino = ws.cell(
            row=fila_destino,
            column=col
        )

        if origen.has_style:
            destino._style = copy(
                origen._style
            )

        destino.number_format = (
            origen.number_format
        )

        if origen.font:
            destino.font = copy(
                origen.font
            )

        if origen.fill:
            destino.fill = copy(
                origen.fill
            )

        if origen.border:
            destino.border = copy(
                origen.border
            )

        if origen.alignment:
            destino.alignment = copy(
                origen.alignment
            )

        if origen.protection:
            destino.protection = copy(
                origen.protection
            )


def generar_excel_resultado(
    resultados
):
    """
    Conserva el libro original y sus otras hojas.

    Para las filas que ya existían:
    - conserva fórmulas existentes;
    - actualiza únicamente celdas normales.

    Para las ventas nuevas:
    - agrega filas al final de la hoja VENTAS;
    - copia el estilo de la última fila existente.
    """
    datos_originales = resultados[
        "bytes_comisiones_original"
    ]

    wb = load_workbook(
        io.BytesIO(
            datos_originales
        ),
        keep_links=True,
    )

    hoja = resultados.get(
        "hoja_comisiones",
        HOJA_COMISIONES
    )

    if hoja not in wb.sheetnames:
        raise ValueError(
            f"No encontré la hoja "
            f"'{hoja}' al preparar "
            f"el Excel final."
        )

    ws = wb[
        hoja
    ]

    final = resultados[
        "final"
    ].copy()

    cantidad_original = int(
        resultados[
            "cantidad_original"
        ]
    )

    # pandas devuelve las columnas en el mismo orden
    # del Excel. Eso es importante porque el archivo
    # real tiene encabezados repetidos como SERIAL
    # y CONCATENAR.
    columnas_excel = [
        col
        for col in final.columns
        if not str(
            col
        ).startswith("__")
    ]

    # La hoja real tiene cabecera en fila 1.
    fila_encabezado = 1
    primera_fila_datos = 2

    max_columnas_escribir = min(
        len(columnas_excel),
        ws.max_column
    )

    # -----------------------------------------------------
    # 1. ACTUALIZAR FILAS EXISTENTES
    # -----------------------------------------------------

    for indice in range(
        min(
            cantidad_original,
            len(final)
        )
    ):
        fila_excel = (
            primera_fila_datos
            + indice
        )

        registro = final.iloc[
            indice
        ]

        for posicion in range(
            max_columnas_escribir
        ):
            columna_df = columnas_excel[
                posicion
            ]

            celda = ws.cell(
                row=fila_excel,
                column=posicion + 1
            )

            # No destruir fórmulas existentes.
            if celda.data_type == "f":
                continue

            valor = registro.get(
                columna_df
            )

            try:
                if pd.isna(valor):
                    valor = None
            except Exception:
                pass

            celda.value = valor

    # -----------------------------------------------------
    # 2. AGREGAR VENTAS NUEVAS
    # -----------------------------------------------------

    fila_ultima_original = (
        primera_fila_datos
        + cantidad_original
        - 1
    )

    if fila_ultima_original < primera_fila_datos:
        fila_ultima_original = (
            primera_fila_datos
        )

    for indice in range(
        cantidad_original,
        len(final)
    ):
        fila_excel = (
            primera_fila_datos
            + indice
        )

        # Copiar formato de la última fila original.
        if ws.max_row >= 2:
            copiar_estilo_fila(
                ws,
                min(
                    fila_ultima_original,
                    ws.max_row
                ),
                fila_excel,
            )

        registro = final.iloc[
            indice
        ]

        for posicion in range(
            max_columnas_escribir
        ):
            columna_df = columnas_excel[
                posicion
            ]

            valor = registro.get(
                columna_df
            )

            try:
                if pd.isna(valor):
                    valor = None
            except Exception:
                pass

            ws.cell(
                row=fila_excel,
                column=posicion + 1
            ).value = valor

    salida = io.BytesIO()

    wb.save(
        salida
    )

    salida.seek(0)

    return salida.getvalue()

